"""
BNIX Webmail - FastAPI Backend
Serves the webmail UI and all API endpoints.
"""
import asyncio
import base64
import binascii
import hashlib
import json
import mimetypes
import os
import re
import secrets
import socket
import ssl
import time
import unicodedata
import uuid
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
from email.header import Header, decode_header, make_header
from email.message import EmailMessage
from email.parser import BytesParser as EmailParser
from email.policy import default as email_default
from email.utils import formataddr, getaddresses
from functools import partial
from pathlib import Path
from typing import Any, AsyncIterator
from urllib.parse import quote

import aioimaplib
import httpx
import uvicorn
from fastapi import FastAPI, File, HTTPException, Request, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, EmailStr, Field

# ─── Config ───────────────────────────────────────────────────────────────────

AUTH_SECRET = os.environ.get(
    "AUTH_SECRET", "An_elNMjOgQouJJCOgPFcChGXNEnXgbDv3E0cQQ6WQM"
)
SESSION_COOKIE = "webmail_session"
SESSION_MAX_AGE = 60 * 60 * 12  # 12 hours
IMAP_TIMEOUT = 60  # seconds
SMTP_TIMEOUT = 30
POOL_TTL = 4 * 60  # 4 minutes

DATA_DIR = os.environ.get("DATA_DIR", "/opt/bnix-webmail/data")
os.makedirs(f"{DATA_DIR}/signatures", exist_ok=True)
os.makedirs(f"{DATA_DIR}/signatures/img", exist_ok=True)
os.makedirs(f"{DATA_DIR}/db", exist_ok=True)

CONTACTS_DB = os.path.join(DATA_DIR, "db", "contacts.db")


def _contacts_init():
    """Init contacts SQLite DB schema once per process."""
    import sqlite3
    with sqlite3.connect(CONTACTS_DB) as db:
        db.execute("""CREATE TABLE IF NOT EXISTS contacts (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            uid        TEXT    UNIQUE NOT NULL,
            account    TEXT    NOT NULL,
            fn         TEXT    NOT NULL DEFAULT '',
            email      TEXT    NOT NULL DEFAULT '',
            phone      TEXT    DEFAULT '',
            organization TEXT  DEFAULT '',
            title      TEXT    DEFAULT '',
            note       TEXT    DEFAULT '',
            photo      TEXT    DEFAULT '',
            created_at TEXT    NOT NULL,
            updated_at TEXT    NOT NULL
        )""")
        db.execute("CREATE TABLE IF NOT EXISTS contacts_migrated (account TEXT PRIMARY KEY)")
        db.commit()


CALENDAR_DB = os.path.join(DATA_DIR, "db", "calendar.db")


def _calendar_init():
    """Init calendar SQLite DB schema once per process."""
    import sqlite3
    with sqlite3.connect(CALENDAR_DB) as db:
        db.execute("""CREATE TABLE IF NOT EXISTS events (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            uid         TEXT    UNIQUE NOT NULL,
            account     TEXT    NOT NULL,
            summary     TEXT    NOT NULL DEFAULT '',
            description TEXT    DEFAULT '',
            location    TEXT    DEFAULT '',
            dtstart     TEXT    NOT NULL DEFAULT '',
            dtend       TEXT    DEFAULT '',
            all_day     INTEGER NOT NULL DEFAULT 0,
            recurrence  TEXT    DEFAULT '',
            attendees   TEXT    DEFAULT '',
            created_at  TEXT    NOT NULL,
            updated_at  TEXT    NOT NULL
        )""")
        db.execute("CREATE TABLE IF NOT EXISTS events_migrated (account TEXT PRIMARY KEY)")
        db.commit()


def _event_row(row: tuple) -> dict:
    return {
        "uid": row[1],
        "summary": row[3] or "",
        "description": row[4] or "",
        "location": row[5] or "",
        "dtstart": row[6] or "",
        "dtend": row[7] or "",
        "allDay": bool(row[8]),
        "recurrence": row[9] or "",
        "attendees": row[10] or "",
    }


_contacts_init()
_calendar_init()

# ─── Labels Database ──────────────────────────────────────────────────────────

LABELS_DB = os.path.join(DATA_DIR, "db", "labels.db")


def _labels_init():
    """Init labels SQLite DB schema once per process."""
    import sqlite3
    with sqlite3.connect(LABELS_DB) as db:
        db.execute("""CREATE TABLE IF NOT EXISTS labels (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            uid        TEXT    UNIQUE NOT NULL,
            account    TEXT    NOT NULL,
            name       TEXT    NOT NULL DEFAULT '',
            color      TEXT    NOT NULL DEFAULT '#6366f1',
            created_at TEXT    NOT NULL,
            updated_at TEXT    NOT NULL
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS message_labels (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            account    TEXT    NOT NULL,
            message_id TEXT    NOT NULL,
            label_uid  TEXT    NOT NULL,
            folder     TEXT    NOT NULL DEFAULT 'INBOX',
            created_at TEXT    NOT NULL,
            UNIQUE(account, message_id, label_uid)
        )""")
        db.execute("CREATE INDEX IF NOT EXISTS idx_ml_account ON message_labels(account)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_ml_msg ON message_labels(account, message_id)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_ml_label ON message_labels(label_uid)")
        db.commit()


def _label_row(row: tuple) -> dict:
    return {
        "uid": row[1],
        "name": row[3] or "",
        "color": row[4] or "#6366f1",
    }


_labels_init()

# ─── Admin Database ──────────────────────────────────────────────────────────

ADMIN_DB = os.path.join(DATA_DIR, "db", "admin.db")
ADMIN_SESSION_COOKIE = "webmail_admin_session"
ADMIN_SESSION_MAX_AGE = 60 * 60 * 8  # 8 hours


def _admin_init():
    """Init admin SQLite DB schema."""
    import sqlite3
    with sqlite3.connect(ADMIN_DB) as db:
        db.execute("""CREATE TABLE IF NOT EXISTS admin_users (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            username   TEXT    UNIQUE NOT NULL,
            password   TEXT    NOT NULL,
            created_at TEXT    NOT NULL,
            updated_at TEXT    NOT NULL
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS domain_aliases (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            alias_domain  TEXT UNIQUE NOT NULL,
            target_domain TEXT NOT NULL DEFAULT '',
            enabled    INTEGER NOT NULL DEFAULT 1,
            created_at TEXT    NOT NULL,
            updated_at TEXT    NOT NULL
        )""")
        # Seed default admin if none exists
        cur = db.execute("SELECT COUNT(*) FROM admin_users")
        if cur.fetchone()[0] == 0:
            import hashlib
            now = datetime.utcnow().isoformat()
            pwd_hash = hashlib.sha256("admin123".encode()).hexdigest()
            db.execute(
                "INSERT INTO admin_users (username, password, created_at, updated_at) VALUES (?,?,?,?)",
                ("admin", pwd_hash, now, now),
            )
        db.commit()


_admin_init()


def _admin_encrypt_session(data: dict) -> str:
    """Encrypt admin session data (reuses main session encryption)."""
    return encrypt_session(data)


def _admin_decrypt_session(token: str | None) -> dict | None:
    """Decrypt admin session data."""
    if not token:
        return None
    try:
        iv_b64, tag_b64, data_b64 = token.split(".")
        from cryptography.hazmat.primitives.ciphers.aead import AESGCM
        aesgcm = AESGCM(_session_key())
        decrypted = aesgcm.decrypt(
            _b64_decode(iv_b64),
            _b64_decode(tag_b64) + _b64_decode(data_b64),
            None,
        )
        data = json.loads(decrypted.decode())
        if not data.get("admin") or not data.get("username"):
            return None
        if time.time() - data.get("createdAt", 0) / 1000.0 > ADMIN_SESSION_MAX_AGE:
            return None
        return data
    except Exception:
        return None


async def require_admin(request: Request) -> dict:
    """Require a valid admin session."""
    token = request.cookies.get(ADMIN_SESSION_COOKIE)
    session = _admin_decrypt_session(token)
    if not session:
        raise HTTPException(401, "Admin authentication required.")
    return session


# Static files directory (served by FastAPI)
STATIC_DIR = Path(__file__).parent / "static"

# ─── IMAP Pool ────────────────────────────────────────────────────────────────

_pool: dict[str, tuple[aioimaplib.IMAP4_SSL, float]] = {}
_pool_lock = asyncio.Lock()
_executor = ThreadPoolExecutor(4)


async def _get_pooled_imap(email: str, password: str, imap_host: str, imap_port: int) -> aioimaplib.IMAP4_SSL:
    """Get or create a pooled IMAP connection."""
    async with _pool_lock:
        entry = _pool.get(email)
        now = time.time()
        if entry:
            client, last_used = entry
            if now - last_used < POOL_TTL:
                # Trust connection is alive — NOOP can corrupt protocol state
                _pool[email] = (client, now)
                return client
            # Stale — evict
            del _pool[email]
            try:
                await client.logout()
            except Exception:
                pass

        client = aioimaplib.IMAP4_SSL(
            host=imap_host,
            port=imap_port,
            timeout=IMAP_TIMEOUT,
        )
        # aioimaplib requires waiting for server greeting before any command
        await client.wait_hello_from_server()
        await client.login(email, password)
        _pool[email] = (client, now)
        return client


async def _evict_imap(email: str):
    """Remove a broken connection from the pool."""
    async with _pool_lock:
        entry = _pool.pop(email, None)
        if entry:
            try:
                await entry[0].logout()
            except Exception:
                pass


async def _evict_pool():
    """Periodic cleanup of stale IMAP connections."""
    async with _pool_lock:
        now = time.time()
        for email in list(_pool.keys()):
            _, last_used = _pool[email]
            if now - last_used > POOL_TTL:
                del _pool[email]


async def with_imap_retry(session: dict, coro_factory):
    """
    Execute an IMAP operation with automatic retry on connection failure.
    coro_factory: a callable that takes (client) and returns a coroutine.
    """
    email = session["email"]
    domain = email.split("@")[1].lower()
    imap_host = session.get("imap_host") or os.environ.get("IMAP_HOST", "").strip() or await _discover_mail_host(domain, "imap")
    imap_port = int(session.get("imap_port") or os.environ.get("IMAP_PORT", "993"))

    for attempt in range(2):
        try:
            client = await _get_pooled_imap(email, session["password"], imap_host, imap_port)
            return await coro_factory(client)
        except (aioimaplib.Abort, ConnectionError, OSError, asyncio.TimeoutError):
            if attempt == 0:
                await _evict_imap(email)
                continue
            raise


async def _get_imap_for_session(session: dict) -> aioimaplib.IMAP4_SSL:
    """Resolve IMAP config for a session."""
    email = session["email"]
    domain = email.split("@")[1].lower()
    imap_host = session.get("imap_host") or os.environ.get("IMAP_HOST", "").strip() or await _discover_mail_host(domain, "imap")
    imap_port = int(session.get("imap_port") or os.environ.get("IMAP_PORT", "993"))
    return await _get_pooled_imap(email, session["password"], imap_host, imap_port)


async def _discover_mail_host(domain: str, service: str = "imap") -> str:
    """
    Auto-discover mail server — Roundcube-style with TCP probing.
    Tries: mail.{domain} → {domain} → MX record → mail.{domain} (fallback)
    For SMTP, probes both port 465 (SMTPS) and 587 (STARTTLS).
    """
    ports = [993] if service == "imap" else [465, 587]

    candidates = [f"mail.{domain}", domain]

    # Add MX record hosts
    try:
        import dns.resolver
        resolver = dns.resolver.Resolver()
        resolver.timeout = 3
        answers = resolver.resolve(domain, "MX")
        mx_hosts = sorted([str(r.exchange).rstrip(".") for r in answers])
        candidates.extend(mx_hosts)
    except Exception:
        pass

    # Always have mail.{domain} as final fallback
    if f"mail.{domain}" not in candidates:
        candidates.append(f"mail.{domain}")

    # Probe each candidate with TCP connect
    for host in candidates:
        for port in ports:
            try:
                _, writer = await asyncio.wait_for(
                    asyncio.open_connection(host, port, ssl=False),
                    timeout=5,
                )
                writer.close()
                await writer.wait_closed()
                # Return host:port as "host" string for SMTP with non-default port
                if service == "smtp" and port != 465:
                    return f"{host}:{port}"
                return host
            except Exception:
                continue

    # Last resort
    return candidates[0]


# ─── Session (AES-256-GCM, matching Next.js) ──────────────────────────────────

def _session_key() -> bytes:
    return hashlib.sha256(AUTH_SECRET.encode()).digest()


def _b64_encode(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).rstrip(b"=").decode()


def _b64_decode(s: str) -> bytes:
    s += "=" * (-len(s) % 4)  # pad to multiple of 4
    return base64.urlsafe_b64decode(s)


def decrypt_session(token: str | None) -> dict | None:
    if not token:
        return None
    try:
        iv_b64, tag_b64, data_b64 = token.split(".")
        from cryptography.hazmat.primitives.ciphers.aead import AESGCM
        aesgcm = AESGCM(_session_key())
        decrypted = aesgcm.decrypt(
            _b64_decode(iv_b64),
            _b64_decode(tag_b64) + _b64_decode(data_b64),
            None,
        )
        session = json.loads(decrypted.decode())
        if not session.get("email") or not session.get("password") or not session.get("createdAt"):
            return None
        if time.time() - session["createdAt"] / 1000.0 > SESSION_MAX_AGE:
            return None
        return session
    except Exception:
        return None


def encrypt_session(session: dict) -> str:
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    import secrets

    aesgcm = AESGCM(_session_key())
    iv = secrets.token_bytes(12)
    encrypted = aesgcm.encrypt(iv, json.dumps(session).encode(), None)
    # Node.js AES-256-GCM: ciphertext = tag (16 bytes) + encrypted
    tag = encrypted[:16]
    data = encrypted[16:]
    return f"{_b64_encode(iv)}.{_b64_encode(tag)}.{_b64_encode(data)}"


# ─── Middleware: session ──────────────────────────────────────────────────────

app = FastAPI(title="BNIX Webmail API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


async def require_session(request: Request) -> dict:
    token = request.cookies.get(SESSION_COOKIE)
    session = decrypt_session(token)
    if not session:
        raise HTTPException(401, "Not authenticated.")
    return session


# ─── Pydantic models ──────────────────────────────────────────────────────────

class Address(BaseModel):
    name: str | None = None
    address: str | None = None

    class Config:
        populate_by_name = True


class MessageSummary(BaseModel):
    uid: int
    messageId: str | None = None
    subject: str
    from_: list[Address] = Field(default_factory=lambda: [], alias="from")
    to: list[Address] = Field(default_factory=list)
    date: str | None = None
    flags: list[str] = Field(default_factory=list)
    seen: bool = False
    flagged: bool = False
    snippet: str = ""


class MessageDetail(BaseModel):
    uid: int
    messageId: str | None = None
    subject: str
    from_: list[Address] = Field(default_factory=lambda: [], alias="from")
    to: list[Address] = Field(default_factory=list)
    date: str | None = None
    flags: list[str] = Field(default_factory=list)
    seen: bool = False
    flagged: bool = False
    snippet: str = ""
    cc: list[Address] = Field(default_factory=list)
    bcc: list[Address] = Field(default_factory=list)
    html: str | None = None
    text: str | None = None
    attachments: list[dict] = Field(default_factory=list)


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _flags_array(flags) -> list[str]:
    if isinstance(flags, set):
        return [str(f) for f in flags]
    if isinstance(flags, (list, tuple)):
        return [str(f) for f in flags]
    return []


def _addresses(raw) -> list[Address]:
    if not isinstance(raw, (list, tuple)):
        return []
    result = []
    for item in raw:
        # aioimaplib envelope returns: (name_bytes, additive_bytes) tuples
        if isinstance(item, (list, tuple)) and len(item) >= 2:
            name_raw = item[0]
            addr_raw = item[1] if len(item) > 1 else None
            name = name_raw.decode("utf-8", errors="replace").strip('"') if name_raw else None
            addr = addr_raw.decode("utf-8", errors="replace") if addr_raw else None
            result.append(Address(name=name or None, address=addr or None))
        elif isinstance(item, bytes):
            result.append(Address(address=item.decode("utf-8", errors="replace")))
        elif isinstance(item, dict):
            result.append(Address(name=item.get("name"), address=item.get("address")))
        elif isinstance(item, str):
            result.append(Address(address=item))
    return result


def _clean_snippet(text: str | None) -> str:
    if not text:
        return ""
    return re.sub(r"\s+", " ", text).strip()[:220]


def _sanitize_html(html: str) -> str:
    """Minimal sanitizer: removes XSS vectors, preserves CSS."""
    out = html
    # Remove dangerous tags
    dangerous_tags = [
        r"<script[\s>][\s\S]*?</script>",
        r"<script[\s\S]*?/>",
        r"<noscript[\s>][\s\S]*?</noscript>",
        r"<iframe[\s>][\s\S]*?</iframe>",
        r"<iframe[\s\S]*?/>",
        r"<object[\s>][\s\S]*?</object>",
        r"<embed[\s\S]*?/?>",
        r"<applet[\s>][\s\S]*?</applet>",
        r"<svg[\s>][\s\S]*?</svg>",
        r"<math[\s>][\s\S]*?</math>",
    ]
    for pattern in dangerous_tags:
        out = re.sub(pattern, "", out, flags=re.IGNORECASE)

    # Scope style content
    def _scope_style(m):
        css = m.group(1)
        css = re.sub(r"expression\s*\([^)]*\)", "", css, flags=re.IGNORECASE)
        css = re.sub(r"behavior\s*:", "", css, flags=re.IGNORECASE)
        css = re.sub(r"-moz-binding\s*:", "", css, flags=re.IGNORECASE)
        css = re.sub(r"url\s*\(\s*['\"]?\s*javascript:", "url(about:blank)", css, flags=re.IGNORECASE)

        def _scope_selector(m) -> str:
            sels = m.group(1) if hasattr(m, "group") else str(m)
            scoped = []
            for s in sels.split(","):
                s = s.strip()
                if not s or s.startswith("@") or re.match(r"^\d+%$", s):
                    scoped.append(s)
                else:
                    scoped.append(f".email-html {s}")
            return ", ".join(scoped)

        css = re.sub(r"([^{}]+)\{", _scope_selector, css)
        return f"<style>{css}</style>"

    out = re.sub(r"<style[^>]*>([\s\S]*?)</style>", _scope_style, out, flags=re.IGNORECASE)

    # Remove event handlers
    out = re.sub(r"\s+on[a-z]+\s*=\s*(?:\"[^\"]*\"|'[^']*'|[^\s>]+)", "", out, flags=re.IGNORECASE)

    # Remove javascript: URLs
    out = re.sub(
        r"\b(href|src|action)\s*=\s*(?:\"javascript:[^\"]*\"|'javascript:[^']*')",
        r'\1="#"',
        out,
        flags=re.IGNORECASE,
    )

    # Fix links
    out = re.sub(r"<a\s", '<a rel="noreferrer noopener" target="_blank" ', out)

    return out


async def _async_parse_email(raw_source: bytes) -> dict:
    """Parse raw email source into structured dict (runs in executor to avoid blocking)."""
    def _parse():
        msg = EmailParser(policy=email_default).parsebytes(raw_source)
        text_parts = []
        html_parts = []
        attachments = []
        cid_urls = {}

        def _walk_payload(part):
            if part.is_multipart():
                return

            content_type = part.get_content_type()
            disp = (part.get_content_disposition() or "").lower()
            cid = part.get("content-id", "").strip("<>") or None
            filename = part.get_filename()
            if not filename:
                name = part.get_param("name", header="content-type")
                if name:
                    filename = name

            payload = part.get_payload(decode=True) or b""
            is_attachment = disp in ("attachment", "inline") or bool(filename) or (
                bool(cid) and content_type.startswith("image/")
            )

            if is_attachment:
                index = len(attachments)
                if cid and payload and content_type.startswith("image/"):
                    cid_urls[cid] = f"data:{content_type};base64,{base64.b64encode(payload).decode()}"
                if not filename and content_type.startswith("image/"):
                    ext = content_type.split("/", 1)[1].split(";", 1)[0] or "image"
                    filename = f"inline-{index + 1}.{ext}"
                attachments.append({
                    "index": index,
                    "filename": filename,
                    "contentType": content_type,
                    "size": len(payload),
                    "cid": cid,
                    "disposition": disp or ("inline" if cid else "attachment"),
                })
                return

            if content_type == "text/plain":
                if payload:
                    text_parts.append(payload.decode(part.get_content_charset() or "utf-8", errors="replace"))
            elif content_type == "text/html":
                if payload:
                    html_parts.append(payload.decode(part.get_content_charset() or "utf-8", errors="replace"))

        def _replace_cid_urls(html: str) -> str:
            for cid, data_url in cid_urls.items():
                escaped = re.escape(cid)
                html = re.sub(rf"cid:{escaped}", data_url, html, flags=re.IGNORECASE)
                html = re.sub(rf"cid:{re.escape(quote(cid))}", data_url, html, flags=re.IGNORECASE)
            return html

        if msg.is_multipart():
            for part in msg.walk():
                _walk_payload(part)
        else:
            _walk_payload(msg)

        html = html_parts[0] if html_parts else None
        if html:
            html = _replace_cid_urls(html)

        return {
            "text": "\n".join(text_parts) or None,
            "html": html,
            "attachments": attachments,
        }

    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(_executor, _parse)


async def _async_get_attachment(raw_source: bytes, index: int) -> dict | None:
    """Return a decoded attachment part by the UI attachment index."""
    def _parse():
        msg = EmailParser(policy=email_default).parsebytes(raw_source)
        attachments = []

        for part in msg.walk() if msg.is_multipart() else [msg]:
            if part.is_multipart():
                continue

            content_type = part.get_content_type()
            disp = (part.get_content_disposition() or "").lower()
            cid = part.get("content-id", "").strip("<>") or None
            filename = part.get_filename()
            if not filename:
                name = part.get_param("name", header="content-type")
                if name:
                    filename = name
            payload = part.get_payload(decode=True) or b""
            is_attachment = disp in ("attachment", "inline") or bool(filename) or (
                bool(cid) and content_type.startswith("image/")
            )
            if not is_attachment:
                continue

            current_index = len(attachments)
            if not filename and content_type.startswith("image/"):
                ext = content_type.split("/", 1)[1].split(";", 1)[0] or "image"
                filename = f"inline-{current_index + 1}.{ext}"
            attachments.append({
                "index": current_index,
                "filename": filename or f"attachment-{current_index + 1}",
                "contentType": content_type,
                "payload": payload,
            })

        for item in attachments:
            if item["index"] == index:
                return item
        return None

    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(_executor, _parse)


async def _fetch_message_source(session: dict, folder: str, uid: int) -> tuple[str | None, bytes | None]:
    async def _do(client):
        select_resp = await client.select(_quote_imap_folder(folder))
        _require_imap_ok(select_resp, f"SELECT {folder}")
        fetch_resp = await client.uid("FETCH", str(uid), "(UID FLAGS INTERNALDATE BODY.PEEK[])")
        _require_imap_ok(fetch_resp, "UID FETCH")
        for meta, source in _iter_fetch_literals(fetch_resp):
            if _fetch_uid(meta, uid) == uid:
                return meta, source
        return None, None

    return await with_imap_retry(session, _do)


def _envelope_summary(msg: Any, snippet: str = "") -> dict:
    flags = _flags_array(getattr(msg, "flags", []) or [])
    envelope = getattr(msg, "envelope", None) or {}
    return {
        "uid": int(getattr(msg, "uid", 0)),
        "messageId": envelope.get("message_id", None),
        "subject": envelope.get("subject", "(No subject)") or "(No subject)",
        "from": _addresses(envelope.get("from", [])),
        "to": _addresses(envelope.get("to", [])),
        "date": _iso_date(envelope.get("date")),
        "flags": flags,
        "seen": "\\Seen" in flags,
        "flagged": "\\Flagged" in flags,
        "snippet": snippet,
    }


def _iso_date(date_val) -> str | None:
    if date_val is None:
        return None
    try:
        from email.utils import parsedate_to_datetime
        return parsedate_to_datetime(date_val).isoformat()
    except Exception:
        return str(date_val)


def _imap_lines(response) -> list[Any]:
    """Return response lines from aioimaplib.Response or older tuple-style data."""
    if hasattr(response, "lines"):
        return list(response.lines or [])
    if isinstance(response, tuple) and len(response) >= 2:
        payload = response[1]
        if isinstance(payload, (list, tuple)):
            return list(payload)
        return [payload] if payload else []
    if isinstance(response, (list, tuple)):
        return list(response)
    return []


def _imap_result(response) -> str | None:
    if hasattr(response, "result"):
        return str(response.result)
    if isinstance(response, tuple) and response:
        first = response[0]
        if isinstance(first, bytes):
            return first.decode("utf-8", errors="replace")
        return str(first)
    return None


def _imap_text(value) -> str:
    if isinstance(value, bytearray):
        value = bytes(value)
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)


def _imap_bytes(value) -> bytes:
    if isinstance(value, bytearray):
        return bytes(value)
    if isinstance(value, bytes):
        return value
    if isinstance(value, str):
        return value.encode()
    return b""


def _require_imap_ok(response, action: str):
    result = _imap_result(response)
    if result and result.upper() != "OK":
        detail = "; ".join(_imap_text(line) for line in _imap_lines(response)[:2])
        raise HTTPException(502, f"IMAP {action} failed: {detail or result}")


SPECIAL_MAILBOXES = {
    "sent": {
        "special": "\\Sent",
        "names": ["Sent", "Sent Messages", "Sent Items", "Sent Mail", "INBOX.Sent", "INBOX/Sent"],
    },
    "archive": {
        "special": "\\Archive",
        "names": ["Archive", "Archives", "INBOX.Archive", "INBOX/Archive"],
    },
    "junk": {
        "special": "\\Junk",
        "names": ["Junk", "Spam", "Bulk Mail", "INBOX.Junk", "INBOX.Spam", "INBOX/Junk", "INBOX/Spam"],
    },
    "trash": {
        "special": "\\Trash",
        "names": ["Trash", "Deleted Messages", "Deleted Items", "Deleted", "Bin", "INBOX.Trash", "INBOX/Trash"],
    },
    "drafts": {
        "special": "\\Drafts",
        "names": ["Drafts", "Draft", "INBOX.Drafts", "INBOX/Drafts"],
    },
}


def _imap_ok(response) -> bool:
    result = _imap_result(response)
    return bool(result and result.upper() == "OK")


def _imap_response_detail(response) -> str:
    return "; ".join(_imap_text(line) for line in _imap_lines(response)[:3]) or (_imap_result(response) or "")


def _fold_mailbox_text(value: str) -> str:
    value = (value or "").replace("Đ", "D").replace("đ", "d")
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return value.casefold().strip()


def _mailbox_leaf(path: str) -> str:
    parts = re.split(r"[./\\]+", path or "")
    return next((p for p in reversed(parts) if p), path or "")


def _canonical_mailbox_role(role: str | None) -> str | None:
    role = (role or "").casefold().strip().lstrip("\\")
    if role == "spam":
        role = "junk"
    return role if role in SPECIAL_MAILBOXES else None


def _role_for_mailbox_name(name: str) -> str | None:
    folded = _fold_mailbox_text(name)
    leaf = _fold_mailbox_text(_mailbox_leaf(name))
    for role, info in SPECIAL_MAILBOXES.items():
        aliases = {_fold_mailbox_text(n) for n in info["names"]}
        aliases.add(role)
        if role == "junk":
            aliases.add("spam")
        if folded in aliases or leaf in aliases:
            return role
    return None


def _special_use_from_flags(flags: list[str]) -> str | None:
    lowered = {flag.casefold(): flag for flag in flags}
    for info in SPECIAL_MAILBOXES.values():
        special = info["special"]
        if special.casefold() in lowered:
            return special
    return None


def _mailbox_exact(path_a: str, path_b: str) -> bool:
    return (path_a or "") == (path_b or "") or _fold_mailbox_text(path_a) == _fold_mailbox_text(path_b)


def _mailbox_info_matches_role(mailbox: dict, role: str) -> bool:
    role = _canonical_mailbox_role(role)
    if not role:
        return False
    special = (mailbox.get("specialUse") or "").casefold()
    if special and special == SPECIAL_MAILBOXES[role]["special"].casefold():
        return True
    return _role_for_mailbox_name(mailbox.get("path") or mailbox.get("name") or "") == role


def _parse_mailbox_list_line(line: Any) -> dict | None:
    line_text = _imap_text(line).strip()
    if not line_text:
        return None
    match = re.match(
        r"^(?:\* LIST\s+)?\((?P<flags>.*?)\)\s+"
        r"(?P<delim>NIL|\"(?:\\.|[^\"])*\"|\S+)\s+"
        r"(?P<name>\"(?:\\.|[^\"])*\"|.+)$",
        line_text,
        flags=re.IGNORECASE,
    )
    if not match:
        return None
    flags = match.group("flags").split()
    delim = _unquote_imap_token(match.group("delim")) or "/"
    name = _unquote_imap_token(match.group("name"))
    if not name:
        return None
    return {"flags": flags, "delimiter": delim, "path": name}


async def _list_mailboxes_for_client(client, include_status: bool = False) -> list[dict]:
    list_resp = await client.list('""', "*")
    _require_imap_ok(list_resp, "LIST")

    mailboxes = []
    for line in _imap_lines(list_resp):
        parsed = _parse_mailbox_list_line(line)
        if not parsed:
            continue

        name = parsed["path"]
        delim = parsed["delimiter"]
        flags = parsed["flags"]
        msgs, unseen = 0, 0

        if include_status:
            try:
                status_resp = await client.status(_quote_imap_folder(name), "(MESSAGES UNSEEN)")
                _require_imap_ok(status_resp, f"STATUS {name}")
                status_str = next(
                    (_imap_text(item) for item in _imap_lines(status_resp) if "MESSAGES" in _imap_text(item)),
                    "",
                )
                msgs_match = re.search(r"MESSAGES\s+(\d+)", status_str)
                unseen_match = re.search(r"UNSEEN\s+(\d+)", status_str)
                msgs = int(msgs_match.group(1)) if msgs_match else 0
                unseen = int(unseen_match.group(1)) if unseen_match else 0
            except Exception:
                msgs, unseen = 0, 0

        mailboxes.append({
            "path": name,
            "name": name,
            "delimiter": delim,
            "specialUse": _special_use_from_flags(flags),
            "total": msgs,
            "unseen": unseen,
            "depth": name.count(delim) if delim and delim in name else 0,
        })

    return mailboxes


async def _create_mailbox_if_missing(client, path: str) -> str:
    create_resp = await client.create(_quote_imap_folder(path))
    if _imap_ok(create_resp):
        try:
            await client.subscribe(_quote_imap_folder(path))
        except Exception:
            pass
        return path

    detail = _imap_response_detail(create_resp)
    if "exist" in detail.casefold() or "already" in detail.casefold():
        return path
    raise HTTPException(502, f"IMAP CREATE {path} failed: {detail or _imap_result(create_resp)}")


async def _ensure_mailbox(client, destination: str, role: str | None = None) -> str:
    destination = (destination or "").strip()
    if not destination and not role:
        raise HTTPException(400, "Destination folder is required.")

    role = _canonical_mailbox_role(role) or _role_for_mailbox_name(destination)
    mailboxes = await _list_mailboxes_for_client(client)

    if destination:
        for mb in mailboxes:
            if _mailbox_exact(mb["path"], destination):
                return mb["path"]

    if role:
        for mb in mailboxes:
            if _mailbox_info_matches_role(mb, role):
                return mb["path"]

        candidates = []
        if destination:
            candidates.append(destination)
        candidates.extend(SPECIAL_MAILBOXES[role]["names"])
        last_error: HTTPException | None = None
        seen: set[str] = set()
        for candidate in candidates:
            folded = _fold_mailbox_text(candidate)
            if not candidate or folded in seen:
                continue
            seen.add(folded)
            try:
                return await _create_mailbox_if_missing(client, candidate)
            except HTTPException as exc:
                last_error = exc
                continue
        if last_error:
            raise last_error

    return await _create_mailbox_if_missing(client, destination)


def _decode_header_value(value) -> str:
    if not value:
        return ""
    try:
        return str(make_header(decode_header(str(value))))
    except Exception:
        return str(value)


def _header_addresses(value) -> list[dict]:
    result = []
    for name, addr in getaddresses([str(value or "")]):
        display_name = _decode_header_value(name).strip() or None
        email_addr = (addr or "").strip() or None
        if display_name or email_addr:
            result.append({"name": display_name, "address": email_addr})
    return result


def _unquote_imap_token(value: str) -> str:
    value = value.strip()
    if value.upper() == "NIL":
        return ""
    if len(value) >= 2 and value[0] == '"' and value[-1] == '"':
        value = value[1:-1].replace(r"\\", "\\").replace(r"\"", '"')
    return _decode_modified_utf7(value)


def _quote_imap_folder(folder: str) -> str:
    """
    Convert folder name from display UTF-8 to IMAP Modified UTF-7,
    then quote it for use in IMAP commands.
    IMAP requires folder names to be either ASCII atoms (7-bit)
    or Modified UTF-7 (RFC 3501) inside double-quotes.
    """
    # Encode UTF-8 → Modified UTF-7
    mutf7 = _encode_modified_utf7(folder)
    # Quote if needed
    needs_quote = any(ord(c) > 127 or c in ("(", ")", "{", " ", "%", "*", '"', "\\") for c in mutf7)
    if not needs_quote:
        return mutf7
    # Escape backslashes and double-quotes for IMAP quoted-string
    escaped = mutf7.replace("\\", "\\\\").replace('"', '\\"')
    return f'"{escaped}"'


def _encode_modified_utf7(s: str) -> str:
    """
    Encode a UTF-8 string to IMAP Modified UTF-7 (RFC 3501).
    ASCII chars are passed through literally.
    Non-ASCII chars are encoded as base64 of UTF-16BE, wrapped in &...-.
    The shift char is '+', and ',' is used instead of '/' in base64.
    """
    import base64
    result = []
    i = 0
    while i < len(s):
        c = ord(s[i])
        if c <= 0x7F:
            # ASCII: pass through literally, escape '&'
            if c == 0x26:  # '&'
                result.append("&-")  # shift out, literal '&', shift back
            else:
                result.append(chr(c))
            i += 1
        else:
            # Non-ASCII: collect run of non-ASCII chars
            start = i
            while i < len(s) and ord(s[i]) > 0x7F:
                i += 1
            chunk = s[start:i]
            # Encode as UTF-16BE, then base64
            utf16 = chunk.encode("utf-16-be")
            b64 = base64.b64encode(utf16).decode("ascii")
            # Modified UTF-7 uses ',' instead of '/' in base64 and strips padding
            b64 = b64.replace("/", ",")
            b64 = b64.rstrip("=")
            result.append(f"&{b64}-")
    return "".join(result)


def _decode_modified_utf7(s: str) -> str:
    """Decode IMAP Modified UTF-7 (RFC 3501) to UTF-8."""
    if "&" not in s:
        return s
    result = []
    i = 0
    while i < len(s):
        amp = s.find("&", i)
        if amp == -1:
            result.append(s[i:])
            break
        result.append(s[i:amp])
        # Find the closing '-' (shift back to ASCII)
        dash = s.find("-", amp + 1)
        if dash == -1:
            # Unclosed shift — treat rest as literal
            result.append(s[amp:])
            break
        encoded = s[amp + 1:dash]
        if not encoded:
            # "&&-" or just "&-" → literal '&'
            result.append("&")
        else:
            # Modified UTF-7: &...- encodes UTF-16BE as base64 (with ',' instead of '+')
            try:
                # Modified UTF-7 uses ',' instead of '/' in base64
                b64 = encoded.replace(",", "/")
                # Pad to multiple of 4
                b64 += "=" * (-len(b64) % 4)
                decoded = base64.b64decode(b64)
                result.append(decoded.decode("utf-16-be"))
            except Exception:
                # Invalid base64 — keep raw
                result.append(s[amp:dash + 1])
        i = dash + 1
    return "".join(result)


def _fetch_uid(meta: str, fallback: int = 0) -> int:
    match = re.search(r"\bUID\s+(\d+)\b", meta)
    return int(match.group(1)) if match else fallback


def _fetch_flags(meta: str) -> list[str]:
    match = re.search(r"\bFLAGS\s+\(([^)]*)\)", meta)
    if not match:
        return []
    return [flag for flag in match.group(1).split() if flag]


def _iter_fetch_literals(response):
    """Yield (FETCH metadata line, literal bytes) pairs from aioimaplib FETCH response."""
    meta = None
    for item in _imap_lines(response):
        text = _imap_text(item)
        if re.match(r"^\d+\s+FETCH\b", text, flags=re.IGNORECASE):
            meta = text
            continue

        if meta is None:
            continue

        literal = _imap_bytes(item)
        if not literal or literal.strip() == b")":
            continue

        yield meta, literal
        meta = None


def _summary_from_header_source(uid: int, meta: str, raw_source: bytes, snippet: str = "") -> dict:
    parsed = EmailParser(policy=email_default).parsebytes(raw_source)
    flags = _fetch_flags(meta)
    return {
        "uid": uid,
        "messageId": parsed.get("Message-ID"),
        "inReplyTo": parsed.get("In-Reply-To"),
        "references": parsed.get("References"),
        "subject": _decode_header_value(parsed.get("Subject")) or "(No subject)",
        "from": _header_addresses(parsed.get("From")),
        "to": _header_addresses(parsed.get("To")),
        "date": _iso_date(parsed.get("Date")),
        "flags": flags,
        "seen": "\\Seen" in flags,
        "flagged": "\\Flagged" in flags,
        "snippet": snippet,
    }


# ─── Routes ───────────────────────────────────────────────────────────────────

@app.get("/health")
async def health():
    return {"ok": True}


# ── Auth ──────────────────────────────────────────────────────────────────────

@app.post("/api/auth/login")
async def login(request: Request, body: dict):
    email = (body.get("email") or "").lower().strip()
    password = body.get("password") or ""
    remember = body.get("remember", True)
    imap_host_override = (body.get("imapHost") or "").strip()
    imap_port_override = body.get("imapPort") or ""
    smtp_host_override = (body.get("smtpHost") or "").strip()
    smtp_port_override = body.get("smtpPort") or ""

    if not email or "@" not in email or not password:
        raise HTTPException(400, "Invalid credentials.")

    # Resolve IMAP server
    domain = email.split("@")[1].lower()

    # Check domain alias
    import sqlite3 as _sqlite3
    try:
        with _sqlite3.connect(ADMIN_DB) as _adb:
            _alias_row = _adb.execute(
                "SELECT target_domain FROM domain_aliases WHERE alias_domain=? AND enabled=1",
                (domain,),
            ).fetchone()
        if _alias_row and _alias_row[0]:
            domain = _alias_row[0].lower()
    except Exception:
        pass

    imap_host = imap_host_override or os.environ.get("IMAP_HOST", "").strip() or await _discover_mail_host(domain, "imap")
    imap_port = int(imap_port_override or os.environ.get("IMAP_PORT", "993"))

    # Resolve SMTP server (for sending)
    smtp_host_raw = smtp_host_override or os.environ.get("SMTP_HOST", "").strip() or await _discover_mail_host(domain, "smtp")
    if ":" in smtp_host_raw and not smtp_host_raw.startswith("["):
        smtp_host, smtp_discovered_port = smtp_host_raw.rsplit(":", 1)
        smtp_port = int(smtp_port_override or smtp_discovered_port)
    else:
        smtp_host = smtp_host_raw
        smtp_port = int(smtp_port_override or os.environ.get("SMTP_PORT", "465"))

    client = aioimaplib.IMAP4_SSL(host=imap_host, port=imap_port, timeout=IMAP_TIMEOUT)
    try:
        await client.wait_hello_from_server()
        await client.login(email, password)
        await client.logout()
    except Exception as e:
        raise HTTPException(401, "Invalid email or password.")

    session = {
        "email": email,
        "password": password,
        "createdAt": int(time.time() * 1000),
        "imap_host": imap_host,
        "imap_port": imap_port,
        "smtp_host": smtp_host,
        "smtp_port": smtp_port,
    }
    # Store server overrides in session if user provided them
    if imap_host_override:
        session["imap_host"] = imap_host_override
    if imap_port_override:
        session["imap_port"] = int(imap_port_override)
    if smtp_host_override:
        session["smtp_host"] = smtp_host_override
    if smtp_port_override:
        session["smtp_port"] = int(smtp_port_override)

    response = JSONResponse({"email": email, "domain": domain})
    response.set_cookie(
        key=SESSION_COOKIE,
        value=encrypt_session(session),
        httponly=True,
        secure=True,
        samesite="lax",
        path="/",
        max_age=SESSION_MAX_AGE if remember else None,
    )
    return response


@app.post("/api/auth/logout")
async def logout(request: Request):
    response = JSONResponse({"ok": True})
    response.delete_cookie(SESSION_COOKIE, path="/")
    return response


@app.get("/api/auth/me")
async def me(request: Request):
    # Public endpoint: returns auth state without throwing 401, so the client
    # can probe the session on page load without polluting the console.
    session = await _load_session(request)
    if not session:
        return {"authenticated": False}
    return {"authenticated": True, "email": session["email"], "domain": session["email"].split("@")[1]}


async def _load_session(request: Request) -> dict | None:
    """Return the decrypted session or None if missing/invalid. Never raises."""
    token = request.cookies.get(SESSION_COOKIE)
    if not token:
        return None
    return decrypt_session(token)


# ── Mailboxes ──────────────────────────────────────────────────────────────────

@app.get("/api/mailboxes")
async def list_mailboxes(request: Request):
    session = await require_session(request)

    async def _do(client):
        return await _list_mailboxes_for_client(client, include_status=True)

    mailboxes = await with_imap_retry(session, _do)
    return JSONResponse({"mailboxes": mailboxes})


@app.post("/api/mailboxes")
async def create_mailbox(request: Request, body: dict):
    session = await require_session(request)
    path = (body.get("path") or "").strip()
    if not path:
        raise HTTPException(400, "Path is required.")

    async def _do(client):
        await _create_mailbox_if_missing(client, path)

    await with_imap_retry(session, _do)
    return JSONResponse({"ok": True, "path": path}, status_code=201)


@app.delete("/api/mailboxes/{path:path}")
async def delete_mailbox(request: Request, path: str):
    """Delete (subscribe + delete) a custom mailbox."""
    session = await require_session(request)
    # Protect main folders
    protected = {"inbox", "trash", "spam", "drafts", "sent", "archive", "inbox/", "trash/", "spam/", "drafts/", "sent/", "archive/"}
    if path.lower() in protected or any(path.lower().startswith(p + "/") for p in protected):
        raise HTTPException(403, "Cannot delete system folder")

    async def _do(client):
        try:
            await client.delete(_quote_imap_folder(path))
        except Exception:
            pass  # IMAP DELETE fails if folder doesn't exist or has messages — ignore

    await with_imap_retry(session, _do)
    return JSONResponse({"ok": True})


# ── Messages ──────────────────────────────────────────────────────────────────

@app.get("/api/messages")
async def list_messages(request: Request):
    session = await require_session(request)
    folder = request.query_params.get("folder", "INBOX")
    limit = min(int(request.query_params.get("limit", "40")), 100)
    offset = max(int(request.query_params.get("offset", "0")), 0)

    async def _do(client):
        select_resp = await client.select(_quote_imap_folder(folder))
        _require_imap_ok(select_resp, f"SELECT {folder}")

        search_resp = await client.uid_search("ALL")
        _require_imap_ok(search_resp, "UID SEARCH")
        uids: list[str] = []
        for item in _imap_lines(search_resp):
            text = _imap_text(item).strip()
            if re.fullmatch(r"[\d\s]+", text):
                uids.extend(text.split())

        total = len(uids)

        if not uids or offset >= total:
            return []

        page_uids = uids[-(offset + limit) : -offset if offset else None]
        if not page_uids:
            page_uids = uids[-limit:]

        uid_set = ",".join(page_uids)
        fetch_resp = await client.uid(
            "FETCH",
            uid_set,
            "(UID FLAGS INTERNALDATE BODY.PEEK[HEADER.FIELDS (SUBJECT FROM TO DATE)])",
        )
        _require_imap_ok(fetch_resp, "UID FETCH")
        return fetch_resp, total

    result = await with_imap_retry(session, _do)
    if not result:
        return JSONResponse({"messages": [], "total": 0, "offset": offset, "limit": limit})

    messages_data, total = result

    messages = []
    for meta, msg_data in _iter_fetch_literals(messages_data):
        try:
            uid = _fetch_uid(meta)
            if uid:
                messages.append(_summary_from_header_source(uid, meta, msg_data))
        except Exception:
            pass

    messages.sort(key=lambda m: m["uid"], reverse=True)
    return JSONResponse({
        "messages": messages,
        "total": total,
        "offset": offset,
        "limit": limit,
    })


def _chunk_messages(data) -> list:
    """Chunk IMAP fetch response into (uid_str, message_data) pairs."""
    result = []
    if not data:
        return result
    if isinstance(data, dict):
        items = list(data.items())
    elif isinstance(data, (list, tuple)):
        items = []
        i = 0
        while i < len(data):
            item = data[i]
            if isinstance(item, (bytes, str)) and re.match(r"\d+.*UID", str(item)):
                uid_str = str(item)
                if i + 1 < len(data):
                    items.append((uid_str, data[i + 1]))
                    i += 2
                    continue
            i += 1
    else:
        items = []
    return items



@app.get("/api/thread")
async def get_thread(request: Request):
    """
    Find all messages belonging to a thread across multiple folders.
    Searches by subject (normalized) in INBOX + Sent folder, then merges.
    Returns summaries sorted oldest→newest.
    """
    session = await require_session(request)
    subject_raw = request.query_params.get("subject", "")
    current_folder = request.query_params.get("folder", "INBOX")
    if not subject_raw:
        raise HTTPException(400, "subject is required")

    # Normalize subject: strip Re:/Fwd: prefixes
    norm_subject = re.sub(r"^(Re|Fwd?|Tr|Fw):\s*", "", subject_raw, flags=re.IGNORECASE).strip()

    # Discover Sent folder name
    async def _find_sent_folder(client):
        for mailbox in await _list_mailboxes_for_client(client):
            if _mailbox_info_matches_role(mailbox, "sent"):
                return mailbox["path"]
        return None

    sent_folder = None
    try:
        sent_folder = await with_imap_retry(session, _find_sent_folder)
    except Exception:
        pass

    # Folders to search: current + Sent (deduplicated)
    folders_to_search = [current_folder]
    if sent_folder and sent_folder.upper() != current_folder.upper():
        folders_to_search.append(sent_folder)

    all_summaries: list[dict] = []

    for folder in folders_to_search:
        async def _search_folder(client, _folder=folder, _subj=norm_subject):
            try:
                sel = await client.select(_quote_imap_folder(_folder), readonly=True)
                _require_imap_ok(sel, f"SELECT {_folder}")
                # IMAP SEARCH by subject text
                search_resp = await client.uid("SEARCH", "SUBJECT", f'"{_subj}"')
                _require_imap_ok(search_resp, "UID SEARCH SUBJECT")
                uids: list[str] = []
                for item in _imap_lines(search_resp):
                    txt = _imap_text(item).strip()
                    if re.fullmatch(r"[\d\s]+", txt):
                        uids.extend(txt.split())
                if not uids:
                    return []
                uid_set = ",".join(uids[-50:])
                fetch_resp = await client.uid(
                    "FETCH", uid_set,
                    "(UID FLAGS INTERNALDATE BODY.PEEK[HEADER.FIELDS (SUBJECT FROM TO DATE MESSAGE-ID IN-REPLY-TO REFERENCES)])"
                )
                _require_imap_ok(fetch_resp, "UID FETCH thread")
                results = []
                for meta, msg_data in _iter_fetch_literals(fetch_resp):
                    try:
                        uid = _fetch_uid(meta)
                        if uid:
                            s = _summary_from_header_source(uid, meta, msg_data)
                            s["folder"] = _folder
                            results.append(s)
                    except Exception:
                        pass
                return results
            except Exception:
                return []

        try:
            msgs = await with_imap_retry(session, _search_folder)
            all_summaries.extend(msgs)
        except Exception:
            pass

    # Deduplicate by messageId, then by uid+folder
    seen_ids: set[str] = set()
    seen_uid_folder: set[tuple] = set()
    deduped = []
    for m in all_summaries:
        mid = (m.get("messageId") or "").strip()
        uf = (m["uid"], m.get("folder", ""))
        if mid and mid in seen_ids:
            continue
        if uf in seen_uid_folder:
            continue
        if mid:
            seen_ids.add(mid)
        seen_uid_folder.add(uf)
        deduped.append(m)

    # Sort oldest → newest by uid (proxy for date within same server)
    deduped.sort(key=lambda m: m["uid"])

    return JSONResponse({"messages": deduped})


@app.get("/api/messages/labels")
async def get_message_labels(request: Request):
    """Get all message-label associations for the current account."""
    session = await require_session(request)
    import sqlite3
    with sqlite3.connect(LABELS_DB) as db:
        rows = db.execute(
            """SELECT ml.message_id, ml.label_uid, l.name, l.color, ml.folder
               FROM message_labels ml
               JOIN labels l ON l.uid = ml.label_uid
               WHERE ml.account = ?
               ORDER BY l.name""",
            (session["email"],)
        ).fetchall()
    result = {}
    for row in rows:
        msg_id = row[0]
        if msg_id not in result:
            result[msg_id] = []
        result[msg_id].append({
            "labelUid": row[1],
            "name": row[2],
            "color": row[3],
            "folder": row[4],
        })
    return JSONResponse(result)


@app.get("/api/messages/{uid}")
async def get_message(request: Request, uid: int):
    session = await require_session(request)
    folder = request.query_params.get("folder", "INBOX")

    async def _do(client):
        select_resp = await client.select(_quote_imap_folder(folder))
        _require_imap_ok(select_resp, f"SELECT {folder}")
        fetch_resp = await client.uid("FETCH", str(uid), "(UID FLAGS INTERNALDATE BODY.PEEK[])")
        _require_imap_ok(fetch_resp, "UID FETCH")
        for meta, source in _iter_fetch_literals(fetch_resp):
            if _fetch_uid(meta, uid) == uid:
                return meta, source
        return None, None

    try:
        meta, source = await asyncio.wait_for(with_imap_retry(session, _do), timeout=60)
    except asyncio.TimeoutError:
        raise HTTPException(504, "Message fetch timed out.")
    if not source:
        raise HTTPException(404, "Message not found.")

    async def _mark_seen(client):
        select_resp = await client.select(_quote_imap_folder(folder))
        _require_imap_ok(select_resp, f"SELECT {folder}")
        store_resp = await client.uid("STORE", str(uid), "+FLAGS", "\\Seen")
        _require_imap_ok(store_resp, "UID STORE")

    try:
        message = EmailParser(policy=email_default).parsebytes(source)
        summary = _summary_from_header_source(uid, meta or "", source)
        summary["cc"] = _header_addresses(message.get("Cc"))
        summary["bcc"] = _header_addresses(message.get("Bcc"))
        summary["html"] = None
        summary["text"] = None
        summary["attachments"] = []

        parsed = await _async_parse_email(source)
        summary["snippet"] = _clean_snippet(parsed.get("text", ""))
        html = parsed.get("html")
        summary["html"] = _sanitize_html(html) if html else None
        summary["text"] = parsed.get("text")
        summary["attachments"] = parsed.get("attachments", [])

        if "\\Seen" not in summary["flags"]:
            try:
                await with_imap_retry(session, _mark_seen)
            except Exception:
                pass

    except asyncio.TimeoutError:
        raise
    except Exception as e:
        raise HTTPException(502, f"Message parse failed: {e}")

    return JSONResponse({"message": summary})


@app.get("/api/messages/{uid}/attachments/{index}")
async def get_attachment(request: Request, uid: int, index: int):
    session = await require_session(request)
    folder = request.query_params.get("folder", "INBOX")
    download = request.query_params.get("download") == "1"

    _, source = await _fetch_message_source(session, folder, uid)
    if not source:
        raise HTTPException(404, "Message not found.")

    attachment = await _async_get_attachment(source, index)
    if not attachment:
        raise HTTPException(404, "Attachment not found.")

    filename = attachment["filename"] or f"attachment-{index + 1}"
    disposition = "attachment" if download else "inline"
    safe_name = quote(filename)
    headers = {
        "Content-Disposition": f"{disposition}; filename*=UTF-8''{safe_name}",
        "Cache-Control": "private, max-age=300",
    }
    return Response(
        content=attachment["payload"],
        media_type=attachment["contentType"] or "application/octet-stream",
        headers=headers,
    )


@app.patch("/api/messages/{uid}/flags")
async def update_flags(request: Request, uid: int, body: dict):
    session = await require_session(request)
    folder = body.get("folder", "INBOX")
    flag = body.get("flag", "\\Seen")
    enabled = bool(body.get("enabled", True))

    async def _do(client):
        select_resp = await client.select(_quote_imap_folder(folder))
        _require_imap_ok(select_resp, f"SELECT {folder}")
        cmd = "+FLAGS" if enabled else "-FLAGS"
        store_resp = await client.uid("STORE", str(uid), cmd, flag)
        _require_imap_ok(store_resp, "UID STORE")

    await with_imap_retry(session, _do)
    return JSONResponse({"ok": True})


@app.post("/api/messages/{uid}/move")
async def move_message(request: Request, uid: int, body: dict):
    session = await require_session(request)
    folder = body.get("folder", "INBOX")
    destination = body.get("destination", "Trash")
    role_value = body.get("role")
    role = _canonical_mailbox_role(role_value if isinstance(role_value, str) else None)

    async def _do(client):
        target = await _ensure_mailbox(client, destination, role=role)
        select_resp = await client.select(_quote_imap_folder(folder))
        _require_imap_ok(select_resp, f"SELECT {folder}")

        copy_resp = await client.uid("COPY", str(uid), _quote_imap_folder(target))
        if not _imap_ok(copy_resp) and "TRYCREATE" in _imap_response_detail(copy_resp).upper():
            target = await _create_mailbox_if_missing(client, target)
            copy_resp = await client.uid("COPY", str(uid), _quote_imap_folder(target))
        _require_imap_ok(copy_resp, "UID COPY")

        store_resp = await client.uid("STORE", str(uid), "+FLAGS", "\\Deleted")
        _require_imap_ok(store_resp, "UID STORE")
        expunge_resp = await client.expunge()
        _require_imap_ok(expunge_resp, "EXPUNGE")
        return target

    target = await with_imap_retry(session, _do)
    return JSONResponse({"ok": True, "destination": target})


@app.delete("/api/messages/{uid}")
async def delete_message(request: Request, uid: int):
    session = await require_session(request)
    folder = request.query_params.get("folder", "INBOX")

    async def _do(client):
        select_resp = await client.select(_quote_imap_folder(folder))
        _require_imap_ok(select_resp, f"SELECT {folder}")
        store_resp = await client.uid("STORE", str(uid), "+FLAGS", "\\Deleted")
        _require_imap_ok(store_resp, "UID STORE")
        expunge_resp = await client.expunge()
        _require_imap_ok(expunge_resp, "EXPUNGE")

    await with_imap_retry(session, _do)
    return JSONResponse({"ok": True})


# ── Send ───────────────────────────────────────────────────────────────────────

@app.post("/api/messages/send")
async def send_message(request: Request, body: dict):
    session = await require_session(request)
    email = session["email"]
    password = session["password"]

    # Parse recipients
    def parse_recipients(value: str) -> list[str]:
        if not value:
            return []
        emails = re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", value, re.IGNORECASE)
        return list(dict.fromkeys(e.lower() for e in emails))

    to_recipients = parse_recipients(body.get("to", ""))
    cc_recipients = parse_recipients(body.get("cc", ""))
    bcc_recipients = parse_recipients(body.get("bcc", ""))
    if not to_recipients:
        raise HTTPException(400, "At least one recipient is required.")

    subject = body.get("subject", "(No subject)")
    text = body.get("text", "")
    html = body.get("html")
    from_name = body.get("fromName", "")
    reply_to = body.get("replyTo", "")

    # Build MIME email with proper RFC 2047/UTF-8 header encoding.
    mime_message = EmailMessage()
    if from_name:
        mime_message["From"] = formataddr((str(Header(from_name, "utf-8")), email))
    else:
        mime_message["From"] = email
    mime_message["To"] = ", ".join(to_recipients)
    if cc_recipients:
        mime_message["Cc"] = ", ".join(cc_recipients)
    mime_message["Subject"] = str(Header(subject, "utf-8"))
    if reply_to:
        mime_message["Reply-To"] = reply_to

    # Threading headers
    in_reply_to = body.get("inReplyTo", "")
    references = body.get("references", "")
    if in_reply_to:
        mime_message["In-Reply-To"] = in_reply_to
        mime_message["References"] = (references + " " + in_reply_to).strip() if references else in_reply_to

    if html:
        mime_message.set_content(text or "", subtype="plain", charset="utf-8")
        mime_message.add_alternative(html, subtype="html", charset="utf-8")
    else:
        mime_message.set_content(text or "", subtype="plain", charset="utf-8")

    # Attachments
    attachments = body.get("attachments", [])
    for att in attachments:
        att_name = att.get("name", "attachment")
        mime_type = att.get("type", "application/octet-stream")
        data_b64 = att.get("data", "")
        # Strip data URL prefix if present (e.g. "data:application/pdf;base64,ABC...")
        if data_b64 and "," in data_b64[:100]:
            data_b64 = data_b64.split(",", 1)[1]
        if data_b64:
            try:
                data_bytes = base64.b64decode(data_b64)
                main, sub = (mime_type.split("/") + ["octet-stream"])[:2]
                mime_message.add_attachment(
                    data_bytes,
                    maintype=main,
                    subtype=sub,
                    filename=att_name,
                )
            except Exception:
                pass  # Skip invalid attachments

    # Send via SMTP
    domain = email.split("@")[1].lower()
    smtp_host_raw = session.get("smtp_host") or os.environ.get("SMTP_HOST", "").strip() or await _discover_mail_host(domain, "smtp")
    # Handle host:port format from discovery (e.g., "mail.domain.com:587")
    if ":" in smtp_host_raw and not smtp_host_raw.startswith("["):
        smtp_host, discovered_port = smtp_host_raw.rsplit(":", 1)
        smtp_port = int(session.get("smtp_port") or discovered_port)
    else:
        smtp_host = smtp_host_raw
        smtp_port = int(session.get("smtp_port") or os.environ.get("SMTP_PORT", "465"))
    smtp_secure = os.environ.get("SMTP_SECURE", "true").lower() != "false"

    import aiosmtplib

    smtp = aiosmtplib.SMTP(
        hostname=smtp_host,
        port=smtp_port,
        use_tls=smtp_secure and smtp_port == 465,
        start_tls=smtp_secure and smtp_port != 465,
        timeout=SMTP_TIMEOUT,
    )
    try:
        await smtp.connect()
        await smtp.login(email, password)
        all_recipients = to_recipients + cc_recipients + bcc_recipients
        print(f"[SMTP] Sending from {email} to {all_recipients} via {smtp_host}:{smtp_port}")
        await smtp.send_message(
            mime_message,
            sender=email,
            recipients=all_recipients,
        )
        await smtp.quit()
        print(f"[SMTP] Send ok — from {email} to {all_recipients}")
    except Exception as e:
        print(f"[SMTP] Failed: {e}")
        raise HTTPException(502, f"SMTP error: {e}")

    # Save copy to Sent folder. Do not fail the SMTP send if IMAP append fails,
    # but surface the warning to the UI so the user is not misled.
    sent_folder = (body.get("sentFolder") or "").strip()
    sent_saved = False
    sent_error = None

    try:
        raw_bytes = mime_message.as_bytes()

        async def _append_sent(client):
            target = await _ensure_mailbox(client, sent_folder or "Sent", role="sent")
            append_resp = await client.append(raw_bytes, _quote_imap_folder(target), flags="\\Seen")
            if not _imap_ok(append_resp) and "TRYCREATE" in _imap_response_detail(append_resp).upper():
                await _create_mailbox_if_missing(client, target)
                append_resp = await client.append(raw_bytes, _quote_imap_folder(target), flags="\\Seen")
            _require_imap_ok(append_resp, f"APPEND {target}")
            print(f"[SENT] Saved to '{target}' ok — from {email}")
            return target

        saved_target = await with_imap_retry(session, _append_sent)
        sent_saved = True
        sent_folder = saved_target
    except Exception as outer_ex:
        sent_error = str(outer_ex)
        print(f"[SENT] Failed to save sent copy: {sent_error}")

    return JSONResponse({"ok": True, "sentSaved": sent_saved, "sentFolder": sent_folder, "sentError": sent_error})


# ── Avatar ─────────────────────────────────────────────────────────────────────

@app.get("/api/avatar")
async def get_avatar(request: Request):
    await require_session(request)
    email = request.query_params.get("email", "").strip().lower()
    if not email or "@" not in email:
        raise HTTPException(400, "Invalid email.")

    gravatar_hash = hashlib.md5(email.encode()).hexdigest()
    gravatar_url = f"https://www.gravatar.com/avatar/{gravatar_hash}?s=128&d=identicon"

    # Try BIMI
    domain = email.split("@")[1].lower()
    bimi_url = None
    try:
        import dns.resolver
        resolver = dns.resolver.Resolver()
        resolver.timeout = 5
        answers = resolver.resolve(f"default._bimi.{domain}", "TXT")
        for rdata in answers:
            txt = "".join(v.decode("utf-8", errors="replace") if isinstance(v, bytes) else str(v) for v in rdata.strings)
            m = re.search(r"[;:]l\s*=\s*([^;]+)", txt, re.IGNORECASE)
            if m:
                url = m.group(1).strip().strip('"')
                if url.startswith("https://"):
                    bimi_url = url
                    break
    except Exception:
        pass

    response = JSONResponse({"bimiUrl": bimi_url, "gravatarUrl": gravatar_url})
    response.headers["Cache-Control"] = "private, max-age=3600"
    return response


# ── Signature ─────────────────────────────────────────────────────────────────

@app.get("/api/settings/signature")
async def get_signature(request: Request):
    session = await require_session(request)
    return JSONResponse({"settings": _read_signature(session["email"])})


@app.put("/api/settings/signature")
async def put_signature(request: Request, body: dict):
    session = await require_session(request)
    return JSONResponse({"settings": _write_signature(session["email"], body)})


def _read_signature(email: str) -> dict:
    key = hashlib.sha256(email.lower().encode()).hexdigest()
    path = f"{DATA_DIR}/signatures/{key}.json"
    defaults = {
        "displayName": "",
        "email": email,
        "organization": "",
        "replyTo": "",
        "bcc": "",
        "defaultEnabled": True,
        "html": "",
        "text": "",
    }
    try:
        with open(path) as f:
            data = json.load(f)
            defaults.update(data)
    except Exception:
        pass
    return defaults


def _write_signature(email: str, data: dict) -> dict:
    key = hashlib.sha256(email.lower().encode()).hexdigest()
    path = f"{DATA_DIR}/signatures/{key}.json"
    os.makedirs(os.path.dirname(path), exist_ok=True)
    defaults = _read_signature(email)
    defaults.update({k: v for k, v in data.items() if k in defaults})
    defaults["email"] = email
    with open(path, "w") as f:
        json.dump(defaults, f, indent=2)
    return defaults


# ── Signature Image Upload ────────────────────────────────────────────────────

# 2 MB hard limit on signature images. The client checks the same limit early,
# and the server remains the source of truth.
_SIG_IMAGE_MAX_BYTES = 2 * 1024 * 1024
_SIG_IMAGE_EXT = {
    "image/png": "png",
    "image/jpeg": "jpg",
    "image/gif": "gif",
    "image/webp": "webp",
}
_SIG_IMAGE_MAGIC = (
    (b"\x89PNG\r\n\x1a\n", {"image/png"}),
    (b"\xff\xd8\xff", {"image/jpeg"}),
    (b"GIF87a", {"image/gif"}),
    (b"GIF89a", {"image/gif"}),
)


def _sig_image_magic_ok(content_type: str, data: bytes) -> bool:
    """Sniff magic bytes so a renamed .exe can't masquerade as a PNG."""
    if not data:
        return False
    if content_type == "image/webp":
        # RIFF....WEBP — need to look 8 bytes in.
        return data[:4] == b"RIFF" and data[8:12] == b"WEBP"
    for prefix, allowed in _SIG_IMAGE_MAGIC:
        if content_type in allowed and data.startswith(prefix):
            return True
    return False


@app.post("/api/settings/signature/image")
async def upload_signature_image(request: Request, file: UploadFile = File(...)):
    """Accept an image for embedding in a signature, store it on disk
    under a per-user directory, and return a same-origin URL the client
    can use as an <img src>."""
    session = await require_session(request)
    content_type = (file.content_type or "").lower().split(";")[0].strip()
    if content_type not in _SIG_IMAGE_EXT:
        raise HTTPException(415, f"Unsupported content type: {content_type}")

    data = await file.read()
    if not data:
        raise HTTPException(400, "Empty file.")
    if len(data) > _SIG_IMAGE_MAX_BYTES:
        raise HTTPException(413, f"File too large (max {_SIG_IMAGE_MAX_BYTES} bytes).")
    if not _sig_image_magic_ok(content_type, data):
        raise HTTPException(400, "File content does not match declared content type.")

    user_key = hashlib.sha256(session["email"].lower().encode()).hexdigest()
    file_id = secrets.token_hex(16)
    ext = _SIG_IMAGE_EXT[content_type]
    user_dir = f"{DATA_DIR}/signatures/img/{user_key}"
    disk_path = f"{user_dir}/{file_id}.{ext}"

    # Write off the event loop — small but unbounded I/O.
    loop = asyncio.get_event_loop()
    def _write() -> str:
        os.makedirs(user_dir, exist_ok=True)
        with open(disk_path, "wb") as f:
            f.write(data)
        return disk_path

    await loop.run_in_executor(_executor, _write)

    # Defense in depth: confirm we wrote inside the user dir.
    safe = Path(user_dir).resolve()
    if not Path(disk_path).resolve().is_relative_to(safe):
        raise HTTPException(500, "Path resolution failed.")

    url = f"/signature-images/img/{user_key}/{file_id}.{ext}"
    return JSONResponse(
        {"url": url, "contentType": content_type, "size": len(data)},
        status_code=201,
    )


# ── DAV Config ────────────────────────────────────────────────────────────────

def _get_dav_config(session: dict) -> dict:
    """Build CalDAV/CardDAV server config from session + env vars."""
    email = session["email"]
    password = session["password"]
    domain = email.split("@")[1].lower()

    use_https = os.environ.get("DAV_SECURE", "true").lower() != "false"
    dav_port = os.environ.get("DAV_PORT", "2080")
    dav_host = os.environ.get("DAV_HOST", "").strip() or f"mail.{domain}"
    scheme = "https" if use_https else "http"
    server_url = f"{scheme}://{dav_host}:{dav_port}"

    return {
        "serverUrl": server_url,
        "username": email,
        "password": password,
    }


async def _dav_request(method: str, url: str, session: dict, body: str | None = None,
                       headers: dict | None = None, retries: int = 2) -> httpx.Response:
    """Make an authenticated WebDAV request with connection-error retry."""
    config = _get_dav_config(session)
    auth = (config["username"], config["password"])
    hdrs = {"Content-Type": "application/xml; charset=utf-8"}
    if headers:
        hdrs.update(headers)

    last_err = None
    for attempt in range(retries + 1):
        try:
            async with httpx.AsyncClient(verify=False, timeout=30) as client:
                resp = await client.request(method, url, auth=auth, headers=hdrs, content=body)
            return resp
        except (httpx.ConnectError, httpx.RemoteProtocolError,
                Exception) as e:
            last_err = e
            if attempt < retries:
                await asyncio.sleep(0.5 * (attempt + 1))

    raise HTTPException(502, f"CardDAV connection failed after {retries + 1} attempts: {last_err}")


def _parse_ics_date(value: str, params: dict | None = None) -> tuple[str, bool]:
    """Parse ICS date/datetime value. Returns (iso_string, all_day)."""
    all_day = (params and params.get("VALUE") == "DATE") or ("T" not in value and len(value) == 8)

    if all_day and len(value) == 8:
        return f"{value[:4]}-{value[4:6]}-{value[6:8]}", True

    m = re.match(r"^(\d{4})(\d{2})(\d{2})T(\d{2})(\d{2})(\d{2})Z?$", value)
    if m:
        y, mo, d, h, mi, s = m.groups()
        is_utc = value.endswith("Z")
        iso = f"{y}-{mo}-{d}T{h}:{mi}:{s}{'Z' if is_utc else ''}"
        return iso, False

    return value, False


def _format_ics_date(iso_date: str, all_day: bool) -> str:
    """Format ISO date to ICS format."""
    if all_day:
        return iso_date.replace("-", "")[:8]
    try:
        dt = datetime.fromisoformat(iso_date.replace("Z", "+00:00"))
        return dt.strftime("%Y%m%dT%H%M%SZ")
    except Exception:
        return iso_date.replace("-", "").replace(":", "").replace("T", "")[:15] + "Z"


def _parse_ics(ics_text: str) -> dict:
    """Parse ICS text into event dict."""
    event = {}
    for line in ics_text.split("\n"):
        line = line.strip()
        colon_idx = line.find(":")
        if colon_idx == -1:
            continue
        key = line[:colon_idx]
        value = line[colon_idx + 1:].strip()

        parts = key.split(";")
        prop_name = parts[0]
        params = {}
        for p in parts[1:]:
            if "=" in p:
                pk, pv = p.split("=", 1)
                params[pk] = pv

        if prop_name == "UID":
            event["uid"] = value
        elif prop_name == "SUMMARY":
            event["summary"] = value
        elif prop_name == "DESCRIPTION":
            event["description"] = value.replace("\\n", "\n").replace("\\,", ",")
        elif prop_name == "LOCATION":
            event["location"] = value
        elif prop_name == "DTSTART":
            date_str, all_day = _parse_ics_date(value, params)
            event["dtstart"] = date_str
            event["allDay"] = all_day
        elif prop_name == "DTEND":
            date_str, _ = _parse_ics_date(value, params)
            event["dtend"] = date_str
        elif prop_name == "RRULE":
            event["recurrence"] = value

    return event


def _build_ics(event: dict, uid: str | None = None) -> str:
    """Build ICS string from event dict."""
    uid_str = uid or event.get("uid") or str(uuid.uuid4())
    now = _format_ics_date(datetime.utcnow().isoformat() + "Z", False)
    all_day = event.get("allDay", False)

    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//BNIX Webmail//EN",
        "BEGIN:VEVENT",
        f"UID:{uid_str}",
        f"DTSTAMP:{now}",
    ]

    if event.get("dtstart"):
        prefix = "DTSTART;VALUE=DATE" if all_day else "DTSTART"
        lines.append(f"{prefix}:{_format_ics_date(event['dtstart'], all_day)}")
    if event.get("dtend"):
        prefix = "DTEND;VALUE=DATE" if all_day else "DTEND"
        lines.append(f"{prefix}:{_format_ics_date(event['dtend'], all_day)}")

    lines.append(f"SUMMARY:{event.get('summary', '')}")

    if event.get("description"):
        lines.append(f"DESCRIPTION:{event['description'].replace(chr(10), '\\n').replace(',', '\\,')}")
    if event.get("location"):
        lines.append(f"LOCATION:{event['location']}")
    if event.get("recurrence"):
        lines.append(f"RRULE:{event['recurrence']}")

    lines.extend(["END:VEVENT", "END:VCALENDAR"])
    return "\r\n".join(lines)


def _parse_vcard(vcard: str) -> dict:
    """Parse vCard string into contact dict."""
    contact = {}

    def get(key: str) -> str | None:
        m = re.search(rf"^{key}[;:](.+)$", vcard, re.MULTILINE | re.IGNORECASE)
        if m:
            return m.group(1).strip().replace("\\n", "\n").replace("\\,", ",")
        return None

    contact["fn"] = get("FN") or ""
    contact["email"] = get("EMAIL") or ""
    if not contact["email"]:
        m = re.search(r"^EMAIL[^:]*:(.+)$", vcard, re.MULTILINE | re.IGNORECASE)
        contact["email"] = m.group(1).strip() if m else ""

    contact["phone"] = get("TEL") or get("TEL;TYPE=CELL") or get("TEL;TYPE=WORK")
    org = get("ORG")
    contact["organization"] = org.replace(";", ", ") if org else None
    contact["title"] = get("TITLE")
    contact["note"] = get("NOTE")

    # Parse PHOTO — base64-encoded inline photo from vCard
    photo_m = re.search(r"^PHOTO[^:]*:(data:[^;]+;base64,)?([A-Za-z0-9+/=\r\n]+)$", vcard, re.MULTILINE | re.IGNORECASE)
    if photo_m:
        contact["photo"] = (photo_m.group(1) or "data:image/jpeg;base64,") + photo_m.group(2).replace("\r", "").replace("\n", "")

    return contact


def _build_vcard(contact: dict, uid: str | None = None) -> str:
    """Build vCard 3.0 string from contact dict."""
    uid_str = uid or contact.get("email") or str(uuid.uuid4())
    fn = contact.get("fn", "")
    name_parts = fn.split(" ")

    lines = [
        "BEGIN:VCARD",
        "VERSION:3.0",
        f"UID:{uid_str}",
        f"FN:{fn}",
        f"N:{';'.join(reversed(name_parts))};;;",
    ]

    if contact.get("email"):
        lines.append(f"EMAIL;TYPE=INTERNET:{contact['email']}")
    if contact.get("phone"):
        lines.append(f"TEL;TYPE=CELL:{contact['phone']}")
    if contact.get("organization"):
        lines.append(f"ORG:{contact['organization'].replace(',', '\\,')}")
    if contact.get("title"):
        lines.append(f"TITLE:{contact['title']}")
    if contact.get("note"):
        lines.append(f"NOTE:{contact['note'].replace(chr(10), '\\n')}")

    lines.append("END:VCARD")
    return "\r\n".join(lines)


def _xml_text(xml: str, tag: str) -> str | None:
    """Extract text content from an XML tag."""
    m = re.search(rf"<{tag}[^>]*>([^<]*)</{tag}>", xml, re.IGNORECASE)
    return m.group(1).strip() if m else None


# ── Calendar (SQLite) ───────────────────────────────────────────────────────

@app.get("/api/calendar")
async def list_calendar_events(request: Request):
    session = await require_session(request)
    account = session["email"]
    start_str = request.query_params.get("start")
    end_str = request.query_params.get("end")

    import sqlite3
    with sqlite3.connect(CALENDAR_DB) as db:
        db.row_factory = sqlite3.Row
        query = "SELECT * FROM events WHERE account=? ORDER BY dtstart"
        params = [account]
        if start_str:
            query += " AND dtstart >= ?"
            params.append(start_str)
        if end_str:
            query += " AND dtstart <= ?"
            params.append(end_str)
        rows = db.execute(query, params).fetchall()

    events = [_event_row(tuple(r)) for r in rows]
    return JSONResponse({"events": events})


@app.post("/api/calendar")
async def create_calendar_event(request: Request, body: dict):
    session = await require_session(request)
    account = session["email"]

    uid = body.get("uid") or str(uuid.uuid4())
    now = datetime.utcnow().isoformat()

    import sqlite3
    with sqlite3.connect(CALENDAR_DB) as db:
        db.execute(
            """INSERT INTO events
               (uid, account, summary, description, location, dtstart, dtend,
                all_day, recurrence, attendees, created_at, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (uid, account,
             body.get("summary", ""), body.get("description", ""),
             body.get("location", ""), body.get("dtstart", ""),
             body.get("dtend", ""),
             1 if body.get("allDay") else 0,
             body.get("recurrence", ""),
             json.dumps(body.get("attendees", [])),
             now, now),
        )
        db.commit()

    return JSONResponse({
        "uid": uid,
        "summary": body.get("summary", ""),
        "description": body.get("description", ""),
        "location": body.get("location", ""),
        "dtstart": body.get("dtstart", ""),
        "dtend": body.get("dtend", ""),
        "allDay": bool(body.get("allDay")),
        "recurrence": body.get("recurrence", ""),
        "attendees": body.get("attendees", []),
    })


@app.put("/api/calendar/{uid}")
async def update_calendar_event(request: Request, uid: str, body: dict):
    session = await require_session(request)
    account = session["email"]
    now = datetime.utcnow().isoformat()

    import sqlite3
    with sqlite3.connect(CALENDAR_DB) as db:
        cur = db.execute(
            """UPDATE events
               SET summary=?, description=?, location=?, dtstart=?, dtend=?,
                   all_day=?, recurrence=?, attendees=?, updated_at=?
               WHERE uid=? AND account=?""",
            (body.get("summary", ""), body.get("description", ""),
             body.get("location", ""), body.get("dtstart", ""),
             body.get("dtend", ""),
             1 if body.get("allDay") else 0,
             body.get("recurrence", ""),
             json.dumps(body.get("attendees", [])),
             now, uid, account),
        )
        db.commit()
        if cur.rowcount == 0:
            raise HTTPException(404, "Event not found")

    return JSONResponse({"ok": True})


@app.delete("/api/calendar/{uid}")
async def delete_calendar_event(request: Request, uid: str):
    session = await require_session(request)
    account = session["email"]

    import sqlite3
    with sqlite3.connect(CALENDAR_DB) as db:
        cur = db.execute("DELETE FROM events WHERE uid=? AND account=?", (uid, account))
        db.commit()
        if cur.rowcount == 0:
            raise HTTPException(404, "Event not found")

    return JSONResponse({"ok": True})


# ── Calendar utilities ──────────────────────────────────────────────────────

@app.post("/api/calendar/import")
async def import_calendar_event(request: Request, body: dict):
    """Parse .ics attachment data and save as calendar event."""
    session = await require_session(request)
    account = session["email"]

    ics_data = body.get("ics", "")
    if not ics_data:
        raise HTTPException(400, "ICS data is required")

    parsed = _parse_ics(ics_data)
    if not parsed.get("summary") and not parsed.get("dtstart"):
        raise HTTPException(400, "Invalid ICS data")

    uid = parsed.get("uid") or str(uuid.uuid4())
    now = datetime.utcnow().isoformat()

    import sqlite3
    with sqlite3.connect(CALENDAR_DB) as db:
        db.execute(
            """INSERT OR IGNORE INTO events
               (uid, account, summary, description, location, dtstart, dtend,
                all_day, recurrence, attendees, created_at, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (uid, account,
             parsed.get("summary", ""), parsed.get("description", ""),
             parsed.get("location", ""), parsed.get("dtstart", ""),
             parsed.get("dtend", ""),
             1 if parsed.get("allDay") else 0,
             parsed.get("recurrence", ""),
             json.dumps([]),
             now, now),
        )
        db.commit()

    return JSONResponse({
        "uid": uid,
        "summary": parsed.get("summary", ""),
        "dtstart": parsed.get("dtstart", ""),
    })


@app.get("/api/calendar/today")
async def calendar_today_events(request: Request):
    """Get today's events for sidebar widget."""
    session = await require_session(request)
    account = session["email"]

    today = datetime.utcnow().date().isoformat()

    import sqlite3
    with sqlite3.connect(CALENDAR_DB) as db:
        db.row_factory = sqlite3.Row
        rows = db.execute(
            "SELECT * FROM events WHERE account=? AND date(dtstart)=? ORDER BY dtstart",
            (account, today),
        ).fetchall()

    return JSONResponse({"events": [_event_row(tuple(r)) for r in rows]})


@app.post("/api/calendar/send-invite")
async def send_calendar_invite(request: Request, body: dict):
    """Build .ics METHOD:REQUEST and send as email invite via SMTP."""
    session = await require_session(request)
    account = session["email"]
    password = session["password"]

    event = body.get("event", {})
    attendee_list = body.get("attendees", [])

    if not attendee_list:
        raise HTTPException(400, "At least one attendee is required")

    uid = event.get("uid") or str(uuid.uuid4())
    now = _format_ics_date(datetime.utcnow().isoformat() + "Z", False)
    all_day = event.get("allDay", False)

    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//BNIX Webmail//EN",
        "METHOD:REQUEST",
        "BEGIN:VEVENT",
        f"UID:{uid}",
        f"DTSTAMP:{now}",
    ]
    if event.get("dtstart"):
        prefix = "DTSTART;VALUE=DATE" if all_day else "DTSTART"
        lines.append(f"{prefix}:{_format_ics_date(event['dtstart'], all_day)}")
    if event.get("dtend"):
        prefix = "DTEND;VALUE=DATE" if all_day else "DTEND"
        lines.append(f"{prefix}:{_format_ics_date(event['dtend'], all_day)}")
    lines.append(f"SUMMARY:{event.get('summary', '')}")
    if event.get("description"):
        lines.append(f"DESCRIPTION:{event['description'].replace(chr(10), '\\n').replace(',', '\\,')}")
    if event.get("location"):
        lines.append(f"LOCATION:{event['location']}")

    account_name = session.get("display_name", "")
    if account_name:
        lines.append(f"ORGANIZER;CN={account_name}:mailto:{account}")
    else:
        lines.append(f"ORGANIZER:mailto:{account}")

    for att in attendee_list:
        att_email = att.get("email", "")
        att_name = att.get("name", "")
        if att_email:
            if att_name:
                lines.append(f"ATTENDEE;CN={att_name}:mailto:{att_email}")
            else:
                lines.append(f"ATTENDEE:mailto:{att_email}")

    lines.extend(["END:VEVENT", "END:VCALENDAR"])
    ics_content = "\r\n".join(lines)

    to_recipients = [att.get("email", "") for att in attendee_list if att.get("email")]

    mime_msg = EmailMessage()
    mime_msg["From"] = formataddr((str(Header(account_name, "utf-8")), account)) if account_name else account
    mime_msg["To"] = ", ".join(to_recipients)
    mime_msg["Subject"] = str(Header(f"Invitation: {event.get('summary', 'Event')}", "utf-8"))
    mime_msg["Content-Type"] = "multipart/mixed"
    mime_msg["MIME-Version"] = "1.0"

    body_text = (
        f"You are invited to: {event.get('summary', '')}\n"
        f"Time: {event.get('dtstart', '')}\n"
        f"Location: {event.get('location', '')}\n\n"
        f"{event.get('description', '')}"
    )
    mime_msg.set_content(body_text, subtype="plain", charset="utf-8")

    ics_bytes = ics_content.encode("utf-8")
    mime_msg.add_attachment(
        ics_bytes,
        maintype="text",
        subtype="calendar",
        filename="invite.ics",
    )

    domain = account.split("@")[1].lower()
    smtp_host_raw = session.get("smtp_host") or os.environ.get("SMTP_HOST", "").strip() or await _discover_mail_host(domain, "smtp")
    if ":" in smtp_host_raw and not smtp_host_raw.startswith("["):
        smtp_host, discovered_port = smtp_host_raw.rsplit(":", 1)
        smtp_port = int(session.get("smtp_port") or discovered_port)
    else:
        smtp_host = smtp_host_raw
        smtp_port = int(session.get("smtp_port") or os.environ.get("SMTP_PORT", "465"))

    use_tls = os.environ.get("SMTP_SECURE", "true").lower() != "false"
    smtp = aiosmtplib.SMTP(timeout=SMTP_TIMEOUT)
    try:
        await smtp.connect()
        if use_tls and smtp.can_starttls:
            await smtp.starttls()
        await smtp.login(account, password)
        await smtp.send_message(mime_msg)
    finally:
        await smtp.quit()

    return JSONResponse({"ok": True, "sent": len(to_recipients)})


# ── Contacts (SQLite) ────────────────────────────────────────────────────────


# ── Contacts (SQLite) ────────────────────────────────────────────────────────

def _contact_row(row: tuple) -> dict:
    return {
        "uid": row[1],
        "fn": row[3] or "",
        "email": row[4] or "",
        "phone": row[5] or "",
        "organization": row[6] or "",
        "title": row[7] or "",
        "note": row[8] or "",
        "photo": row[9] or "",
    }


@app.get("/api/contacts")
async def list_contacts(request: Request):
    session = await require_session(request)
    account = session["email"]

    import sqlite3
    with sqlite3.connect(CONTACTS_DB) as db:
        db.row_factory = sqlite3.Row
        rows = db.execute(
            "SELECT * FROM contacts WHERE account=? ORDER BY fn",
            (account,),
        ).fetchall()

    return JSONResponse({"contacts": [_contact_row(tuple(r)) for r in rows]})


@app.get("/api/contacts/export")
async def export_contacts(request: Request):
    session = await require_session(request)
    account = session["email"]

    import sqlite3
    with sqlite3.connect(CONTACTS_DB) as db:
        db.row_factory = sqlite3.Row
        rows = db.execute(
            "SELECT fn, email, phone, organization, title, note FROM contacts WHERE account=? ORDER BY fn",
            (account,),
        ).fetchall()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from starlette.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # Header style
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2563EB")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(bottom=thin)

    headers = ["Name", "Email", "Phone", "Organization", "Title", "Note"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    for row in rows:
        ws.append([row["fn"] or "", row["email"] or "", row["phone"] or "",
                   row["organization"] or "", row["title"] or "", row["note"] or ""])

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 40

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"contacts_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/contacts")
async def create_contact(request: Request, body: dict):
    session = await require_session(request)
    account = session["email"]

    uid = body.get("uid") or str(uuid.uuid4())
    now = datetime.utcnow().isoformat()

    import sqlite3
    try:
        with sqlite3.connect(CONTACTS_DB) as db:
            # Check if this account already has a contact with the same email
            email = body.get("email", "").strip()
            if email:
                existing = db.execute(
                    "SELECT uid FROM contacts WHERE account=? AND LOWER(email)=LOWER(?)",
                    (account, email),
                ).fetchone()
                if existing:
                    raise HTTPException(409, "Contact with this email already exists")

            db.execute(
                """INSERT INTO contacts
                   (uid, account, fn, email, phone, organization, title, note, photo, created_at, updated_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (uid, account,
                 body.get("fn", ""), email,
                 body.get("phone", ""), body.get("organization", ""),
                 body.get("title", ""), body.get("note", ""),
                 body.get("photo", ""),
                 now, now),
            )
            db.commit()
    except HTTPException:
        raise
    except sqlite3.IntegrityError:
        raise HTTPException(409, "Contact with this email already exists")

    return JSONResponse({
        "uid": uid,
        "fn": body.get("fn", ""),
        "email": body.get("email", ""),
        "phone": body.get("phone", ""),
        "organization": body.get("organization", ""),
        "title": body.get("title", ""),
        "note": body.get("note", ""),
    })


@app.put("/api/contacts/{uid}")
async def update_contact(request: Request, uid: str, body: dict):
    session = await require_session(request)
    account = session["email"]

    now = datetime.utcnow().isoformat()
    import sqlite3
    with sqlite3.connect(CONTACTS_DB) as db:
        cur = db.execute(
            """UPDATE contacts
               SET fn=?, email=?, phone=?, organization=?, title=?, note=?, photo=?, updated_at=?
               WHERE uid=? AND account=?""",
            (body.get("fn", ""), body.get("email", ""),
             body.get("phone", ""), body.get("organization", ""),
             body.get("title", ""), body.get("note", ""),
             body.get("photo", ""), now, uid, account),
        )
        db.commit()
        if cur.rowcount == 0:
            raise HTTPException(404, "Contact not found")

    return JSONResponse({"ok": True})


@app.delete("/api/contacts/{uid}")
async def delete_contact(request: Request, uid: str):
    session = await require_session(request)
    account = session["email"]

    import sqlite3
    with sqlite3.connect(CONTACTS_DB) as db:
        cur = db.execute("DELETE FROM contacts WHERE uid=? AND account=?", (uid, account))
        db.commit()
        if cur.rowcount == 0:
            raise HTTPException(404, "Contact not found")

    return JSONResponse({"ok": True})


# ─── Admin API ────────────────────────────────────────────────────────────────

@app.post("/api/admin/login")
async def admin_login(request: Request, body: dict):
    username = (body.get("username") or "").strip()
    password = body.get("password") or ""
    if not username or not password:
        raise HTTPException(400, "Username and password required.")

    import hashlib
    pwd_hash = hashlib.sha256(password.encode()).hexdigest()

    import sqlite3
    with sqlite3.connect(ADMIN_DB) as db:
        row = db.execute(
            "SELECT id, username FROM admin_users WHERE username=? AND password=?",
            (username, pwd_hash),
        ).fetchone()

    if not row:
        raise HTTPException(401, "Invalid credentials.")

    session_data = {
        "admin": True,
        "username": row[1],
        "createdAt": int(time.time() * 1000),
    }
    response = JSONResponse({"ok": True, "username": row[1]})
    response.set_cookie(
        key=ADMIN_SESSION_COOKIE,
        value=_admin_encrypt_session(session_data),
        httponly=True,
        secure=True,
        samesite="lax",
        path="/",
        max_age=ADMIN_SESSION_MAX_AGE,
    )
    return response


@app.post("/api/admin/logout")
async def admin_logout():
    response = JSONResponse({"ok": True})
    response.delete_cookie(ADMIN_SESSION_COOKIE, path="/")
    return response


@app.get("/api/admin/me")
async def admin_me(request: Request):
    token = request.cookies.get(ADMIN_SESSION_COOKIE)
    session = _admin_decrypt_session(token)
    if not session:
        return {"authenticated": False}
    return {"authenticated": True, "username": session["username"]}


@app.put("/api/admin/password")
async def admin_change_password(request: Request, body: dict):
    admin = await require_admin(request)
    old_password = body.get("oldPassword") or ""
    new_password = body.get("newPassword") or ""
    if not old_password or not new_password:
        raise HTTPException(400, "Old and new passwords required.")
    if len(new_password) < 6:
        raise HTTPException(400, "Password must be at least 6 characters.")

    import hashlib
    old_hash = hashlib.sha256(old_password.encode()).hexdigest()
    new_hash = hashlib.sha256(new_password.encode()).hexdigest()

    import sqlite3
    with sqlite3.connect(ADMIN_DB) as db:
        row = db.execute(
            "SELECT id FROM admin_users WHERE username=? AND password=?",
            (admin["username"], old_hash),
        ).fetchone()
        if not row:
            raise HTTPException(401, "Current password is incorrect.")
        now = datetime.utcnow().isoformat()
        db.execute(
            "UPDATE admin_users SET password=?, updated_at=? WHERE username=?",
            (new_hash, now, admin["username"]),
        )
        db.commit()

    return JSONResponse({"ok": True})


@app.get("/api/admin/domains")
async def admin_list_domains(request: Request):
    await require_admin(request)
    import sqlite3
    with sqlite3.connect(ADMIN_DB) as db:
        db.row_factory = sqlite3.Row
        rows = db.execute("SELECT * FROM domain_aliases ORDER BY alias_domain").fetchall()
    return JSONResponse({
        "domains": [
            {
                "id": r["id"],
                "aliasDomain": r["alias_domain"],
                "targetDomain": r["target_domain"],
                "enabled": bool(r["enabled"]),
                "createdAt": r["created_at"],
            }
            for r in rows
        ]
    })


@app.post("/api/admin/domains")
async def admin_add_domain(request: Request, body: dict):
    await require_admin(request)
    alias_domain = (body.get("aliasDomain") or "").lower().strip()
    target_domain = (body.get("targetDomain") or "").lower().strip()
    if not alias_domain or not target_domain:
        raise HTTPException(400, "Alias domain and target domain required.")

    now = datetime.utcnow().isoformat()
    import sqlite3
    try:
        with sqlite3.connect(ADMIN_DB) as db:
            db.execute(
                """INSERT INTO domain_aliases (alias_domain, target_domain, enabled, created_at, updated_at)
                   VALUES (?, ?, 1, ?, ?)""",
                (alias_domain, target_domain, now, now),
            )
            db.commit()
    except sqlite3.IntegrityError:
        raise HTTPException(409, "Domain alias already exists.")

    return JSONResponse({"ok": True, "aliasDomain": alias_domain, "targetDomain": target_domain})


@app.put("/api/admin/domains/{domain_id}")
async def admin_update_domain(request: Request, domain_id: int, body: dict):
    await require_admin(request)
    import sqlite3
    now = datetime.utcnow().isoformat()
    with sqlite3.connect(ADMIN_DB) as db:
        cur = db.execute(
            """UPDATE domain_aliases
               SET target_domain=?, enabled=?, updated_at=?
               WHERE id=?""",
            (body.get("targetDomain", ""), 1 if body.get("enabled", True) else 0, now, domain_id),
        )
        db.commit()
        if cur.rowcount == 0:
            raise HTTPException(404, "Domain alias not found.")
    return JSONResponse({"ok": True})


@app.delete("/api/admin/domains/{domain_id}")
async def admin_delete_domain(request: Request, domain_id: int):
    await require_admin(request)
    import sqlite3
    with sqlite3.connect(ADMIN_DB) as db:
        cur = db.execute("DELETE FROM domain_aliases WHERE id=?", (domain_id,))
        db.commit()
        if cur.rowcount == 0:
            raise HTTPException(404, "Domain alias not found.")
    return JSONResponse({"ok": True})


@app.get("/api/admin/domains/lookup/{domain}")
async def admin_lookup_domain(domain: str):
    """Public endpoint: lookup domain alias for login redirect."""
    domain = domain.lower().strip()
    import sqlite3
    with sqlite3.connect(ADMIN_DB) as db:
        row = db.execute(
            "SELECT target_domain FROM domain_aliases WHERE alias_domain=? AND enabled=1",
            (domain,),
        ).fetchone()
    if row and row[0]:
        return JSONResponse({"alias": True, "targetDomain": row[0]})
    return JSONResponse({"alias": False})


# ─── Static Files & SPA ──────────────────────────────────────────────────────

# Per-user signature images live outside STATIC_DIR (under DATA_DIR), so we
# serve them with a dedicated route that:
#   1. Refuses path-traversal attempts (resolved target must stay inside the
#      signatures directory).
#   2. Sets `Cache-Control: private, immutable` — the URL embeds a random hex
#      so `immutable` is honest, and `private` keeps shared proxies from
#      caching user content.
_SIG_SERVE_ROOT = Path(f"{DATA_DIR}/signatures").resolve()


@app.get("/signature-images/{path:path}")
@app.get("/assets/signatures/{path:path}")
async def serve_signature_image(path: str):
    target = (_SIG_SERVE_ROOT / path).resolve()
    if _SIG_SERVE_ROOT not in target.parents or not target.is_file():
        raise HTTPException(404, "Not found.")
    media, _ = mimetypes.guess_type(str(target))
    return FileResponse(
        target,
        media_type=media or "application/octet-stream",
        headers={"Cache-Control": "private, max-age=86400, immutable"},
    )


# Mount static assets (CSS, JS, images). Keep this after the signature image
# route so /assets/signatures/... is not swallowed by the generic /assets mount.
if STATIC_DIR.exists():
    app.mount("/assets", StaticFiles(directory=str(STATIC_DIR / "assets")), name="static-assets")
    app.mount("/brand", StaticFiles(directory=str(STATIC_DIR / "brand")), name="static-brand")


@app.get("/admin")
async def serve_admin():
    """Serve the admin page."""
    admin_path = STATIC_DIR / "admin.html"
    if admin_path.exists():
        return FileResponse(str(admin_path))
    raise HTTPException(404, "Admin page not found.")


@app.get("/")
async def serve_index():
    """Serve the main SPA page."""
    index_path = STATIC_DIR / "index.html"
    if index_path.exists():
        return FileResponse(str(index_path))
    return HTMLResponse("<h1>BNIX Webmail</h1><p>Static files not found. Run build first.</p>")


@app.get("/favicon.png")
async def serve_favicon():
    """Serve favicon."""
    fav = STATIC_DIR / "brand" / "bnix-favicon.png"
    if fav.exists():
        return FileResponse(str(fav), media_type="image/png")
    raise HTTPException(404)


# ─── Startup ─────────────────────────────────────────────────────────────────

@app.get("/api/labels")
async def list_labels(request: Request):
    session = await require_session(request)
    import sqlite3
    with sqlite3.connect(LABELS_DB) as db:
        rows = db.execute(
            "SELECT * FROM labels WHERE account = ? ORDER BY name",
            (session["email"],)
        ).fetchall()
    return JSONResponse([_label_row(r) for r in rows])


@app.post("/api/labels")
async def create_label(request: Request):
    session = await require_session(request)
    body = await request.json()
    name = (body.get("name") or "").strip()
    color = body.get("color") or "#6366f1"
    if not name:
        raise HTTPException(400, "Label name is required")
    uid = str(uuid.uuid4())
    now = datetime.utcnow().isoformat()
    import sqlite3
    with sqlite3.connect(LABELS_DB) as db:
        db.execute(
            "INSERT INTO labels (uid, account, name, color, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)",
            (uid, session["email"], name, color, now, now)
        )
        db.commit()
    return JSONResponse({"uid": uid, "name": name, "color": color})


@app.put("/api/labels/{label_uid}")
async def update_label(request: Request, label_uid: str):
    session = await require_session(request)
    body = await request.json()
    import sqlite3
    with sqlite3.connect(LABELS_DB) as db:
        row = db.execute(
            "SELECT * FROM labels WHERE uid = ? AND account = ?",
            (label_uid, session["email"])
        ).fetchone()
        if not row:
            raise HTTPException(404, "Label not found")
        name = (body.get("name") or row[3]).strip()
        color = body.get("color") or row[4]
        now = datetime.utcnow().isoformat()
        db.execute(
            "UPDATE labels SET name = ?, color = ?, updated_at = ? WHERE uid = ?",
            (name, color, now, label_uid)
        )
        db.commit()
    return JSONResponse({"uid": label_uid, "name": name, "color": color})


@app.delete("/api/labels/{label_uid}")
async def delete_label(request: Request, label_uid: str):
    session = await require_session(request)
    import sqlite3
    with sqlite3.connect(LABELS_DB) as db:
        db.execute(
            "DELETE FROM labels WHERE uid = ? AND account = ?",
            (label_uid, session["email"])
        )
        db.execute(
            "DELETE FROM message_labels WHERE label_uid = ? AND account = ?",
            (label_uid, session["email"])
        )
        db.commit()
    return JSONResponse({"ok": True})


@app.post("/api/messages/{uid}/labels")
async def add_message_label(request: Request, uid: int):
    session = await require_session(request)
    body = await request.json()
    label_uid = body.get("labelUid")
    folder = body.get("folder", "INBOX")
    if not label_uid:
        raise HTTPException(400, "labelUid is required")
    now = datetime.utcnow().isoformat()
    # Use messageId if available, otherwise use uid+folder as identifier
    message_key = body.get("messageId") or f"{folder}:{uid}"
    import sqlite3
    with sqlite3.connect(LABELS_DB) as db:
        try:
            db.execute(
                "INSERT INTO message_labels (account, message_id, label_uid, folder, created_at) VALUES (?, ?, ?, ?, ?)",
                (session["email"], message_key, label_uid, folder, now)
            )
            db.commit()
        except sqlite3.IntegrityError:
            pass  # Already assigned
    return JSONResponse({"ok": True})


@app.delete("/api/messages/{uid}/labels/{label_uid}")
async def remove_message_label(request: Request, uid: int, label_uid: str):
    session = await require_session(request)
    folder = request.query_params.get("folder", "INBOX")
    message_key = request.query_params.get("messageId") or f"{folder}:{uid}"
    import sqlite3
    with sqlite3.connect(LABELS_DB) as db:
        db.execute(
            "DELETE FROM message_labels WHERE account = ? AND message_id = ? AND label_uid = ?",
            (session["email"], message_key, label_uid)
        )
        db.commit()
    return JSONResponse({"ok": True})


@app.on_event("startup")
async def startup():
    asyncio.create_task(_pool_cleanup_loop())


async def _pool_cleanup_loop():
    while True:
        await asyncio.sleep(60)
        await _evict_pool()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8000"))
    host = os.environ.get("HOST", "127.0.0.1")
    uvicorn.run(app, host=host, port=port, log_level="info")
